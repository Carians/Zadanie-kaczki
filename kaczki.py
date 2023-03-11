# Twoje zadanie.py: Dane jest N kaczuszek gumowych, które należy ułożyć w rzędzie. Każda kaczuszka ma określoną wysokość
# i szerokość. Potrzebujemy ułożyć kaczuszki w taki sposób, aby suma wysokości wykorzystanych kaczuszek była jak
# największa, a ich sumaryczna szerokość nie przekroczyła maksymalnej szerokości rzędu.
#
# Wejście: W pierwszym wierszu standardowego wejścia znajdują się dwie liczby całkowite N, M (1 <= N,
# M <= 50) pooddzielane pojedynczymi odstępami oznaczające odpowiednio: liczbę dostępnych kaczuszek oraz maksymalną
# szerokość rzędu. W każdym kolejnym N wierszu znajdują się dwie liczby całkowite w, s (1 <= w, s <=9) oddzielone
# pojedynczym odstępem oznaczające wysokość (w) i szerokość (s) kaczuszki Wyjście: Twój program powinien wypisać na
# standardowe wyjście maksymalną dostępną sumę wysokości użytych kaczuszek do ustawienia ich w rzędzie.
#
# Wyjście: Twój program powinien wypisać na standardowe wyjście maksymalną dostępną sumę wysokości użytych kaczuszek
# do ustawienia ich w rzędzie. W oknie poniżej wpisz sam wynik. Nie sprawdzamy twojego kodu. Możesz próbować do
# skutku, tyle razy ile potrzebujesz – do 23:59 we wtorek 11.04.

# N - liczba kaczuszek
# M - maksymalna szerokosc rzędu

def pobierz_dane_xlsx():
    import openpyxl
    wb = openpyxl.load_workbook('zadanie-rekrutacyjne.xlsx')
    sheet = wb.active
    N = sheet['A1'].value
    M = sheet['B1'].value
    wysokosci = []
    szerokosci = []
    for i in range(2, N + 2):
        wysokosci.append(sheet.cell(row=i, column=1).value)
        szerokosci.append(sheet.cell(row=i, column=2).value)
    return N, M, wysokosci, szerokosci

def Main():
    N = 20 # Liczba kaczuszek
    M = 50 # Maksymalna szerokosc rzędu
    wysokosci = [3, 5, 7, 2, 8, 4, 6, 9, 1, 5, 7, 8, 3, 4, 6, 2, 1, 5, 8, 9]  # Najwieksze
    szerokosci = [2, 4, 6, 1, 5, 3, 7, 9, 2, 4, 6, 8, 3, 5, 7, 1, 2, 4, 6, 8]  # Nie może przekroczyć 50

    # W przypadku gdy dane są w pliku xlsx odkomentować poniższą linijkę. Uwaga - Potrzeba zainstalować bibliotekę openpyxl
    #N, M, wysokosci, szerokosci = pobierz_dane_xlsx()

    x, x2 = 0, 0
    for i in range(0, N):
        x += wysokosci[i]
        x2 += szerokosci[i]
    #print('Wysokości: '+str(x), 'Szerokości: '+str(x2))
    #print(maksymalna_suma_wysokości(N, M, wysokosci, szerokosci))
    print(optymalna_maksymalna_suma_wysokości_wykorzystanych_kaczuszek(M, wysokosci, szerokosci))

def polacz_kaczki(wysokosci, szerokosci):
    kaczki = []
    for i in range(len(wysokosci)):
        kaczki.append([wysokosci[i], szerokosci[i]])
    return kaczki


def optymalna_maksymalna_suma_wysokości_wykorzystanych_kaczuszek(M, wysokosci, szerokosci):
    # Polaczenie wysokosci i szerokosci w jedna liste
    kaczki = polacz_kaczki(wysokosci, szerokosci)
    suma_wysokosci = 0
    suma_szerokosci = 0
    najwieksza_kaczki = []
    ratio = 0

    # Sortowanie kaczek po stosunku wysokosci do szerokosci
    for i in range(len(kaczki)):
        temp_ratio = kaczki[i][0] / kaczki[i][1]
        ratio = temp_ratio
        najwieksza_kaczki.append([kaczki[i], "ratio: " + str(ratio)])
    najwieksza_kaczki.sort(key=lambda x: x[1], reverse=True)

    # Dodawanie kaczek do rzędu
    for i in range(len(kaczki)):
        if suma_szerokosci + najwieksza_kaczki[i][0][1] <= M:
            suma_szerokosci += najwieksza_kaczki[i][0][1]
            suma_wysokosci += najwieksza_kaczki[i][0][0]

    return suma_wysokosci



if __name__ == "__main__":
    Main()

# 65